import openpyxl
from pathlib import Path
data_path = Path(__file__).parent / "data" / "news.xlsx"
import sys


class CustomExcelLibrary:
    @staticmethod
    def write_data_to_excel(excel_path, sheet_name, data_lists):
        """
        Writes data from multiple lists into an Excel file.
        
        :param sheet_name:
        :param excel_path: Path to the Excel file.
        :param data_lists: List of lists containing data to write.
        """
        book = openpyxl.load_workbook(excel_path)
        # sheet = book.active
        sheet = book[sheet_name]
        for row_idx, data_list in enumerate(data_lists, start=1):
            for col_idx, value in enumerate(data_list, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        book.save(excel_path)

    @staticmethod
    def count_keyword(excel_path, keyword):
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_path)

        # Initialize counter
        keyword_count = 0

        # Iterate through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Iterate through all cells in the sheet
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and keyword in str(cell.value):
                        keyword_count += 1
        return keyword_count

    @staticmethod
    def find_most_common_keywords(excel_path, num_keywords=100, exclude_words=None):
        """
        Finds the most frequently appearing keywords in an Excel file.

        :param excel_path: Path to the Excel file
        :param num_keywords: Number of top keywords to return (default: 20)
        :param exclude_words: Optional list of common words to exclude (e.g., 'the', 'and', 'a')
        :return: List of tuples containing (keyword, count) sorted by count in descending order
        """
        # Default exclude list if none provided
        if exclude_words is None:
            exclude_words = [
                "the", "and", "a", "to", "of", "in", "is", "that", "it", "for",
                "on", "with", "as", "at", "this", "by", "be", "or", "an", "are",
                "was", "were", "from", "have", "has", "had", "not", "but", "what",
                "all", "when", "who", "which", "will", "there", "can", "more", "no",
                "if", "so", "their", "would", "about", "up", "out", "them", "then",
                "some", "these", "his", "her", "they", "could", "into", "only", "one",
                "been", "other", "do", "you", "your", "my", "we", "us", "our", "me"
            ]


        # Load the workbook
        workbook = openpyxl.load_workbook(excel_path)

        # Initialize word counter dictionary
        word_counts = {}

        # Iterate through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Iterate through all cells in the sheet
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # Split cell text into words
                        words = str(cell.value).lower().split()

                        # Count each word
                        for word in words:
                            # Remove any punctuation
                            word = word.strip('.,!?:;()"\'')

                            # Skip empty strings or excluded words
                            if word and word not in exclude_words:
                                if word in word_counts:
                                    word_counts[word] += 1
                                else:
                                    word_counts[word] = 1

        # Sort words by count in descending order and get the top N
        if not word_counts:
            return []

        # Sort by count (descending) and get top num_keywords
        top_keywords = sorted(word_counts.items(), key=lambda x: x[1], reverse=True)[:num_keywords]
        return top_keywords
    @staticmethod
    def generate_keyword_chart(top_keywords, chart_title="Top Keywords", output_path=None):
        """
        Generates a bar chart of the top keywords using matplotlib.

        :param top_keywords: List of tuples (keyword, count) from find_most_common_keywords
        :param chart_title: Title for the chart
        :param output_path: Path to save the chart image (optional)
        """
        try:
            import matplotlib.pyplot as plt

            # Unpack keywords and counts
            keywords = [item[0] for item in top_keywords]
            counts = [item[1] for item in top_keywords]

            # Create figure with appropriate size based on number of keywords
            plt.figure(figsize=(10, max(6, len(keywords) * 0.3)))

            # Create horizontal bar chart (easier to read for text labels)
            bars = plt.barh(keywords, counts, color='skyblue')

            # Add count labels on the bars
            for bar in bars:
                width = bar.get_width()
                plt.text(width + 0.3, bar.get_y() + bar.get_height() / 2,
                         f'{width:.0f}', ha='left', va='center')

            # Add labels and title
            plt.xlabel('Occurrences')
            plt.ylabel('Keywords')
            plt.title(chart_title)
            plt.tight_layout()

            # Save the chart if output path is provided
            if output_path:
                plt.savefig(output_path)
                print(f"Chart saved to {output_path}")

            # Display the chart
            plt.show()
        except ImportError:
            print("Matplotlib is not installed. Please install it to generate charts.")

    @staticmethod
    def export_keywords_to_excel_with_chart(top_keywords, output_excel_path, sheet_name="cnn"):
        """
        Exports the top keywords to a new Excel file with a chart.

        :param top_keywords: List of tuples (keyword, count) from find_most_common_keywords
        :param output_excel_path: Path to save the new Excel file
        :param sheet_name: Name of the sheet for the data and chart
        """
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference

        # Create a new workbook and get active sheet
        # wb = Workbook()
        # ws = wb.active
        wb = openpyxl.load_workbook(output_excel_path)
        ws = wb.active
        ws.title = sheet_name

        # Add headers
        ws['A1'] = "Keyword"
        ws['B1'] = "Occurrences"

        # Add data
        for i, (keyword, count) in enumerate(top_keywords, start=2):
            ws.cell(row=i, column=1, value=keyword)
            ws.cell(row=i, column=2, value=count)

        # Create chart
        chart = BarChart()
        chart.type = "bar"  # Horizontal bar chart
        chart.title = "Top Keywords by Frequency"
        chart.y_axis.title = "Keywords"
        chart.x_axis.title = "Occurrences"

        # Add data to chart
        data = Reference(ws, min_col=2, min_row=1, max_row=len(top_keywords) + 1, max_col=2)
        categories = Reference(ws, min_col=1, min_row=2, max_row=len(top_keywords) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Add chart to worksheet (position it below the data)
        ws.add_chart(chart, f"D2")

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # Save the workbook
        wb.save(output_excel_path)
        return f"Excel file with chart created at: {output_excel_path}"

    @staticmethod
    def save_keywords_to_existing_excel(excel_path, top_keywords, sheet_name, add_chart=True):
        """
        Saves the top keywords to a new worksheet in an existing Excel file with formatting.

        :param excel_path: Path to the existing Excel file
        :param top_keywords: List of tuples (keyword, count) from find_most_common_keywords
        :param sheet_name: Name of the new worksheet to create (or use if exists)
        :param add_chart: Whether to add a bar chart to visualize the data
        :return: Status message
        """
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.chart import BarChart, Reference, AreaChart
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            print("Debug message", sheet_name, excel_path)
            # Load the existing workbook
            try:
                workbook = openpyxl.load_workbook(excel_path)
            except FileNotFoundError:
                return f"Error: Excel file not found at {excel_path}"

            # Check if the sheet already exists
            # if sheet_name in workbook.sheetnames:
            #     # Use the existing sheet
            #     worksheet = workbook[sheet_name]
            #     # Clear existing content
            #     for row in worksheet.rows:
            #         for cell in row:
            #             cell.value = None
            # else:
            # Create a new worksheet
            worksheet = workbook.create_sheet(title=sheet_name)
            print("Debug message", sheet_name, excel_path)

            # Define styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Add title
            title_cell = worksheet['A1']
            title_cell.value = "Keyword Analysis Results"
            title_cell.font = Font(bold=True, size=14)
            worksheet.merge_cells('A1:B1')
            title_cell.alignment = Alignment(horizontal='center')

            # Add headers
            worksheet['A2'] = "Keyword"
            worksheet['B2'] = "Occurrences"

            # Apply header styling
            for col in range(1, 3):
                cell = worksheet.cell(row=2, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            # Add data
            for i, (keyword, count) in enumerate(top_keywords, start=3):
                # Add keyword
                keyword_cell = worksheet.cell(row=i, column=1, value=keyword)
                keyword_cell.border = border

                # Add count
                count_cell = worksheet.cell(row=i, column=2, value=count)
                count_cell.border = border
                count_cell.alignment = Alignment(horizontal='center')

                # Add alternating row colors
                if i % 2 == 0:
                    for col in range(1, 3):
                        worksheet.cell(row=i, column=col).fill = PatternFill(
                            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                        )

            # Add a chart if requested
            if add_chart:
                # Create chart
                chart = AreaChart()
                # chart.type = "bar"  # Horizontal bar chart
                chart.title = "Top Keywords by Frequency"
                chart.y_axis.title = "Keywords"
                chart.x_axis.title = "Occurrences"
                chart.style = 10  # A clean chart style

                # Define the data range for the chart
                data = Reference(worksheet, min_col=2, min_row=2, max_row=len(top_keywords) + 2, max_col=2)
                categories = Reference(worksheet, min_col=1, min_row=3, max_row=len(top_keywords) + 2)

                # Add data to chart
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)

                # Add chart to worksheet - position it to the right of the data
                chart_cell = f"D3"
                worksheet.add_chart(chart, chart_cell)

            # Auto-adjust column widths
            for col in range(1, 3):
                max_length = 0
                column_letter = get_column_letter(col)
                for row in range(1, len(top_keywords) + 3):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 4
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook
            workbook.save(excel_path)


            return f"Keywords saved to worksheet '{sheet_name}' in {excel_path} with formatting"

        except Exception as e:
            return f"Error saving keywords to Excel: {str(e)}"


    def get_account(self, excel_path, role):
        book = openpyxl.load_workbook(excel_path)
        sheet = book.active
        role = self.check_role(role)
        row = role + 1
        row_data = []
        for cell in sheet[row]:
            row_data.append(cell.value)
        return row_data

    @staticmethod
    def get_all_rows_except_first(excel_path, user):
        book = openpyxl.load_workbook(excel_path)
        sheet = book.active
        all_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if user in str(row):
                all_rows.append(list(row))
        return all_rows[0]

    @staticmethod
    def check_role(value):
        switch_dict = {
            "admin": 1,
            "hr": 2,
            "employee": 3
        }
        return switch_dict.get(value, 1)

    @staticmethod
    def get_account_by_uid(excel_path, uid):
        book = openpyxl.load_workbook(excel_path)
        sheet = book.active
        row = str(uid).split("_", 2)
        row = int(row[1])
        row_data = []
        for cell in sheet[row]:
            row_data.append(cell.value)
        return row_data

    @staticmethod
    def split_list(input_list, size):
        """
        Splits a single list into multiple smaller lists of a given size.
    
        :param input_list: List to be split
        :param size: Size of each smaller list
        :return: List of smaller lists
        """
        return [input_list[i:i + int(size)] for i in range(0, len(input_list), int(size))]
